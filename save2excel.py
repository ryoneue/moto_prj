import openpyxl

def search_col(sheet, time):
    # 時間から列番号を取得
    target_hour = time
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = header.index(target_hour)
    col_index = col_index + 1 # indexが0から始まるため調整
    return col_index

def search_row(sheet, row_data):
    # 日付と設備名
    row_index = None
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == row_data[0] and row[1].value == row_data[1]:
            row_index = row[0].row
            break
    return row_index

def make_excel_file(excel_file):
    # ファイルが存在する場合は既存のデータを読み込み、存在しない場合は新規作成
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # 既存のシートを取得または新規作成
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name, index=0)  # index=0で先頭に作成する
        sheet.append(header)
    return workbook, sheet

def write_data_to_excel(data, excel_file, sheet_name, header):
    # Excelファイルを取得
    workbook, sheet = make_excel_file(excel_file)


    # データを書き込む
    for date, machine_data in data.items():
        for machine, data in machine_data.items():
            time_data, count_data = [[i, j] for i, j in data.items()][0]
            row_data = [date, machine]
            time_data = time_data + "時"
            row_data.append(count_data)

            row_index = search_row(sheet, row_data)
            time_col_index = search_col(sheet, time_data)
            if row_index:
                # 既存の行を更新
                col_index = [1, 2, time_col_index]
                for i, value in enumerate(row_data):
                    sheet.cell(row=row_index, column=col_index[i], value=value)
            else:
                # 新たに行を作成
                col_index = [1, 2, time_col_index]
                row_index = sheet.max_row + 1 # 行を一つ追加
                for i, value in enumerate(row_data):
                    sheet.cell(row=row_index, column=col_index[i], value=value)

    # ファイルを保存
    workbook.save(excel_file)


if __name__ == "__main__":
    from html_check import html_check

    url = "http://127.0.0.1:8000/"
    page = html_check(url)
    machine_data = page.check_machine(machine_num=3)
    data = machine_data

    # Excelファイルのパスを指定
    excel_file = 'sample.xlsx'
    sheet_name = 'Sheet1'
    header = ['日付', '設備名', '0時','1時','2時','3時','4時','5時','6時','7時','8時','9時', '10時', '11時', '12時', '13時', '14時', '15時', '16時', '17時', '18時', '19時', '20時', '21時', '22時', '23時']

    write_data_to_excel(data, excel_file, sheet_name, header)
